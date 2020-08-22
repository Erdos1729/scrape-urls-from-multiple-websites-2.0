from bs4 import BeautifulSoup
from urllib.request import urlopen
import re
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
import urllib
from fake_useragent import UserAgent
import openpyxl
import http.client
import feedparser
import csv
import datetime
import os
import pandas as pd

def read_csv(path = './input_file/all_urls.csv'):
    with open(path, encoding='utf-8') as f:
        fobj = csv.DictReader(f)

        for row in fobj:
            yield row

getALLlinks = []
Links = read_csv()
for id, link_i in enumerate(Links):
    # print(link_i["\ufeffurls"])
    getALLlinks.append(link_i["\ufeffurls"])

wb = openpyxl.Workbook()
sheet = wb.active
# sheet.title = 'Output'

#Add titles in the first row of each column
sheet.cell(row=1, column=1).value='id'
sheet.cell(row=1, column=2).value='url'
sheet.cell(row=1, column=3).value='title'
sheet.cell(row=1, column=4).value='corrected url'
sheet.cell(row=1, column=5).value='source'
sheet.cell(row=1, column=6).value='extract_date'

wb.save('./output_file/export/scraped_pr_links.xlsx')


excelcounterrow = 2

processedlink = 0
skippedlink = 0
id = []

for lop in range(len(getALLlinks)):
    ua = UserAgent()
    req = Request(getALLlinks[lop])
    req.add_header('User-Agent', ua.random)
    processedlink = processedlink + 1

    print("Processing link no." + str(processedlink) + ": " + getALLlinks[lop])

    wbopen = openpyxl.load_workbook('./output_file/export/scraped_pr_links.xlsx')
    sheetopen = wbopen.active

    try:
        html_page = urlopen(req).read()
    #except (TimeoutError, ConnectionResetError, Exception, URLError, HTTPError, socket.timeout, http.client.RemoteDisconnected,http.client.HTTPException) as e:

    except (TimeoutError, ConnectionResetError, Exception, URLError, HTTPError, http.client.RemoteDisconnected, http.client.HTTPException, ConnectionError) as e:
        skippedlink = skippedlink + 1
        #html_page = urlopen(req)
        print('Unable to open link no.' + str(processedlink) + ": " + getALLlinks[lop])

    except:
        skippedlink = skippedlink + 1
        #html_page = urlopen(req)
        print('Unable to open link no.' + str(processedlink) + ": " + getALLlinks[lop])

    else:
        soup = BeautifulSoup(html_page, features="html.parser")
        LinkYMLIndex = getALLlinks[lop]

        flag = 0

        for link in soup.findAll('a', href=True):
            # skip useless links

            if link['href'] == '' or link['href'].startswith('#'):
                continue

            # initialize the link
            thisLink = {
                'url': link['href'],
                'title': link.string,
            }
            actuallink = getALLlinks[lop]
            slashcounter = 0
            indexslash = 0

            if not actuallink.endswith("/"):
                actuallink = actuallink + "/"

            while slashcounter < 3:
                if(actuallink[indexslash] == "/"):
                    slashcounter = slashcounter + 1
                indexslash = indexslash + 1

            PLink = actuallink[:indexslash - 1]

            ModLink = thisLink['url'].strip()
            NewLink = ''

            if ModLink.startswith("/"):
                NewLink = PLink + ModLink

            else:
                NewLink = ModLink

            if thisLink['title'] is None:
                # check for text inside the link
                if len(link.contents):
                    thisLink['title'] = ' '.join(link.stripped_strings)
            if thisLink['title'] is None:
                # if there's *still* no title (empty tag), skip it
                continue
            # convert to something immutable for storage

            sheetopen.cell(row=excelcounterrow, column=2).value = thisLink['url'].strip().replace('','')
            sheetopen.cell(row=excelcounterrow, column=3).value = thisLink['title'].strip().replace('','')
            sheetopen.cell(row=excelcounterrow, column=4).value = NewLink.replace('','')
            sheetopen.cell(row=excelcounterrow, column=5).value = getALLlinks[lop]
            sheetopen.cell(row=excelcounterrow, column=6).value = datetime.date.today()
            excelcounterrow = excelcounterrow+1
            flag = 1

        if flag == 0:
            d = feedparser.parse(LinkYMLIndex)
            for rss in d.entries:
               linkrss = rss.link
               linktitle = rss.title

               actuallink = getALLlinks[lop]
               slashcounter = 0
               indexslash = 0
               while slashcounter < 3:
                   if (actuallink[indexslash] == "/"):
                       slashcounter = slashcounter + 1
                   indexslash = indexslash + 1
               PLink = actuallink[:indexslash - 1]

               ModLink = linkrss
               NewLink = ''

               if ModLink.startswith("/"):
                   NewLink = PLink + ModLink

               else:
                   NewLink = ModLink

               sheetopen.cell(row=excelcounterrow, column=2).value = linkrss
               sheetopen.cell(row=excelcounterrow, column=3).value = linktitle
               sheetopen.cell(row=excelcounterrow, column=4).value = NewLink
               sheetopen.cell(row=excelcounterrow, column=5).value = getALLlinks[lop]
               sheetopen.cell(row=excelcounterrow, column=6).value = datetime.date.today()
               excelcounterrow = excelcounterrow + 1

    sheet.title = 'Output'
    wbopen.save('./output_file/export/scraped_pr_links.xlsx')

wbopen1 = openpyxl.load_workbook('./output_file/export/scraped_pr_links.xlsx')
ws = wbopen1.get_sheet_by_name('Sheet')  # get sheet by name
sheetopen1 = wbopen1.active
# print(len(sheetopen1['url']))

for i in range(1, len(sheetopen1['url']) + 1):

    if i == 1:
        sheetopen1.cell(row=i, column=1).value = 'id'
    elif i == len(sheetopen1['url']) + 1:
        sheetopen1.cell(row=i, column=1).value = ''
    else:
        sheetopen1.cell(row=i, column=1).value = i - 1

    values = [ws["C" + str(i)].value, ws["D" + str(i)].value]  # collect the data
    values = [str(l) for l in values]
    # print(values)
    sheetopen1.cell(row=i, column=7).value = ' '.join(values) # concat column C and D

sheetopen1.cell(row=1, column=7).value = 'concat'
wbopen1.save('./output_file/export/scraped_pr_links.xlsx')

print("\nTotal Processed Links: " + str(processedlink))
print("Total Unprocessed Links: " + str(skippedlink))

def load_data(name1, name2):
    df, df1 = pd.read_excel(name1), pd.read_excel(name2)
    return df, df1

filename = './output_file/export/scraped_pr_links.xlsx'
filename1 = './output_file/database/allextract_merged.xlsx'
filename2 = './output_file/Final_output.csv'

if os.path.exists(filename1):
    print("Identifying latest links.....")

    # Read an excel with two sheets into two dataframes

    df, df1 = load_data(filename, filename1)

    lookup = []
    for i in [str(l) for l in df['concat']]:
        if i in [str(x) for x in df1['concat']]:
            lookup.append('True')
        else: lookup.append('False')

    df2 = df
    df2['lookup'] = lookup
    df2.to_csv(filename2, index=False)

    df2['lookup'] = [str(m) for m in df2['lookup']]
    # print(df[df2['lookup'] == 'False'])

    df1 = df1.append(df[df2['lookup'] == 'False'])

    # database_update = pd.merge(df2, df1, how = 'left')
    # df1 = database_update
    df1.to_excel(filename1, index=False)
    # print(len(df1['concat']))

else:
    wbopen1.save(filename1)
    wbopen1.save(filename2)